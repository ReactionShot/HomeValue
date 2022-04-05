#! python3
# Scrapes websites to keep track of your property price.

import time
import re
from datetime import date
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl import Workbook
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.firefox.service import Service


def zillow_estimate(address):
    print("Getting Zillow Property Estimate...")
    address = address.replace(' ', '-')
    url = 'https://www.zillow.com/homes/'+address+'_rb'
    driver.get(url)
    time.sleep(3)
    print("Zillow Estimate -", driver.find_element(by=By.XPATH, value='//h3[@class="Text-c11n-8-62-5__sc-aiai24-0 StyledHeading-c11n-8-62-5__sc-ktujwe-0 gzNGZJ"]').text)
    return driver.find_element(by=By.XPATH, value='//h3[@class="Text-c11n-8-62-5__sc-aiai24-0 StyledHeading-c11n-8-62-5__sc-ktujwe-0 gzNGZJ"]').text


def redfin_estimate(address):
    print("Getting Redfin Property Estimate...")
    url = 'https://www.redfin.com/what-is-my-home-worth'
    driver.get(url)
    time.sleep(2)
    driver.find_element(by=By.XPATH, value='/html/body/div/div[8]/div[2]/div/div/div[1]/div[2]/div/form/div/div/input').send_keys(address)
    time.sleep(2)
    driver.find_element(by=By.XPATH, value='/html/body/div/div[8]/div[2]/div/div/div[1]/div[2]/div/form/div/div/input').send_keys(Keys.ENTER)
    time.sleep(3)
    print("Redfin Estimate -", driver.find_element(by=By.XPATH, value='//div[@class="price font-size-larger"]').text)
    return driver.find_element(by=By.XPATH, value='//div[@class="price font-size-larger"]').text


def trulia_estimate(address):
    print("Getting Trulia Property Estimate...")
    url = 'https://www.trulia.com/'
    driver.get(url)
    driver.find_element(by=By.XPATH, value='//input[@id="banner-search"]').click()
    time.sleep(2)
    driver.find_element(by=By.XPATH, value='//input[@id="banner-search"]').send_keys(address)
    time.sleep(2)
    driver.find_element(by=By.XPATH, value='//input[@id="banner-search"]').send_keys(Keys.ENTER)
    time.sleep(3)
    print("Trulia Estimate -", driver.find_element(by=By.XPATH, value='//*[@id="main-content"]/div[2]/div[1]/div/div[1]/div/div/div[2]/div/h3/div').text)
    return driver.find_element(by=By.XPATH, value='//*[@id="main-content"]/div[2]/div[1]/div/div[1]/div/div/div[2]/div/h3/div').text


def homes_estimate(address):
    print("Getting Homes Property Estimate...")
    url = 'https://www.homes.com/what-is-my-home-worth/'
    driver.get(url)
    time.sleep(3)
    driver.find_element(by=By.XPATH, value='//div[@class="relative w-full"]//input[@placeholder="Search by Address"]').send_keys(address)
    time.sleep(3)
    driver.find_element(by=By.XPATH, value='//div[@class="relative w-full"]//input[@placeholder="Search by Address"]').send_keys(Keys.ENTER)
    time.sleep(3)
    split_home_value = driver.find_element(by=By.XPATH, value='//h3[@class="mb-1/4 font-size-xxl font-weight-regular"]').text.split()
    print("Homes Estimate - " + split_home_value[1])
    return split_home_value[1]


if __name__ == '__main__':
    date_today = date.today().strftime('%m/%d/%Y')
    property_street = input('Enter property street address: ')
    property_city = input('Enter property city: ')
    property_state = input('Enter property state: ')
    property_zip = input('Enter property zip code: ')
    address_format = property_street + ', ' + property_city + ', ' + property_state + ', ' + property_zip
    zillow_address_format = property_street + '-' + property_city + ',-' + property_state + '-' + property_zip
    property_fname = property_street.split(" ", 1)[1] + '.xlsx'

    print('Launching Firefox...')
    options = Options()
    s = Service('geckodriver.exe')
    #options.add_argument("--headless")
    driver = webdriver.Firefox(service=s,options=options)
    driver.implicitly_wait(10)
    driver.maximize_window()

    # Clean Estimates
    zillow_value = 0
    redfin_value = 0
    trulia_value = 0
    homes_value = 0

    zillow_value = zillow_estimate(zillow_address_format)
    redfin_value = redfin_estimate(address_format)
    trulia_value = trulia_estimate(address_format)
    homes_value = homes_estimate(address_format)

    driver.close()
    driver.quit()

    try:
        wb = load_workbook(filename=property_fname)
        ws = wb.active
    except FileNotFoundError:
        print("File not found, creating new file...")
        wb = Workbook()
        ws = wb.active
        ws.title = 'Home Estimate'
        ws.column_dimensions['A'].width = 11

        # Set Zillow values
        ws['B1'] = 'Zillow'
        ws.column_dimensions['B'].width = 11
        zillow_cell = ws.cell(row=1, column=2)
        zillow_cell.font = Font(bold=True)
        zillow_cell.alignment = Alignment(horizontal='center')

        # Set Redfin values
        ws['C1'] = 'Redfin'
        ws.column_dimensions['C'].width = 11
        redfin_cell = ws.cell(row=1, column=3)
        redfin_cell.font = Font(bold=True)
        redfin_cell.alignment = Alignment(horizontal='center')

        # Set Trulia values
        ws['D1'] = 'Trulia'
        ws.column_dimensions['D'].width = 11
        trulia_cell = ws.cell(row=1, column=4)
        trulia_cell.font = Font(bold=True)
        trulia_cell.alignment = Alignment(horizontal='center')

        # Set Homes values
        ws['E1'] = 'Homes'
        ws.column_dimensions['E'].width = 11
        homes_cell = ws.cell(row=1, column=5)
        homes_cell.font = Font(bold=True)
        homes_cell.alignment = Alignment(horizontal='center')

        # Set Average value
        ws['F1'] = 'Average'
        ws.column_dimensions['F'].width = 11
        average_cell = ws.cell(row=1, column=6)
        average_cell.font = Font(bold=True)
        average_cell.alignment = Alignment(horizontal='center')

        zillow_value = int(re.sub('\D', '', zillow_value))
        redfin_value = int(re.sub('\D', '', redfin_value))
        trulia_value = int(re.sub('\D', '', trulia_value))
        homes_value = int(re.sub('\D', '', homes_value))

        print("Adding property values for", date_today)
        new_row = ws.max_row + 1
        ws.cell(row=new_row, column=1).value = date_today
        ws.cell(row=new_row, column=2).number_format = '$#,##0.00'
        ws.cell(row=new_row, column=3).number_format = '$#,##0.00'
        ws.cell(row=new_row, column=4).number_format = '$#,##0.00'
        ws.cell(row=new_row, column=5).number_format = '$#,##0.00'
        ws.cell(row=new_row, column=6).number_format = '$#,##0.00'
        ws.cell(row=new_row, column=2).value = zillow_value
        ws.cell(row=new_row, column=3).value = redfin_value
        ws.cell(row=new_row, column=4).value = trulia_value
        ws.cell(row=new_row, column=5).value = homes_value
        ws.cell(row=new_row, column=6).value = (homes_value + trulia_value + redfin_value + zillow_value) / 4
        wb.save(property_fname)
        print('Changes have been saved to', property_fname)

    date_cell = ws.cell(row=ws.max_row, column=1)

    if date_cell.value != date_today:
        print("Adding property values for", date_today)
        new_row = ws.max_row + 1
        ws.cell(row=new_row, column=1).value = date_today
        ws.cell(row=new_row, column=2).number_format = '$#,##0.00'
        ws.cell(row=new_row, column=3).number_format = '$#,##0.00'
        ws.cell(row=new_row, column=4).number_format = '$#,##0.00'
        ws.cell(row=new_row, column=5).number_format = '$#,##0.00'
        ws.cell(row=new_row, column=6).number_format = '$#,##0.00'
        ws.cell(row=new_row, column=2).value = zillow_value
        ws.cell(row=new_row, column=3).value = redfin_value
        ws.cell(row=new_row, column=4).value = trulia_value
        ws.cell(row=new_row, column=5).value = homes_value
        ws.cell(row=new_row, column=6).value = (homes_value + trulia_value + redfin_value + zillow_value) / 4
        wb.save(property_fname)
        print('Changes have been saved to', property_fname)
    else:
        print("The script has already run today")
